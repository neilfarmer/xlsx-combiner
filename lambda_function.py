import boto3
import io
import openpyxl
import os

BUCKET = os.environ["BUCKET"] or "nxf25-garbage-bucket"
NEW_DATA = os.environ["NEW_DATA_FILE"] or "aws_finances.xlsx"  # New CUR data
TEMPLATE = os.environ["TEMPLATE_FILE"] or "template.xlsx"      # Custom XLSX
DEST = os.environ["DESTINATION_FILE"] or "final_report.xlsx"      # Destination filename

NEW_DATA_WORKSHEET_NAME = os.environ["NEW_DATA_WORKSHEET_NAME"] or "costs"   # CUR Worksheet name
TEMPLATED_WORKSHEET_NAME = os.environ["TEMPLATED_WORKSHEET_NAME"] or "data"   # Target Worksheet name

s3 = boto3.client('s3')

def lambda_handler(event, context):

    new_data_file = s3.get_object(Bucket=BUCKET, Key=NEW_DATA)
    template_file = s3.get_object(Bucket=BUCKET, Key=TEMPLATE)

    new_data = new_data_file['Body'].read()
    template = template_file['Body'].read()

    new_data_workbook = openpyxl.load_workbook(io.BytesIO(new_data))
    templated_workbook = openpyxl.load_workbook(io.BytesIO(template))


    new_data_worksheet = new_data_workbook[NEW_DATA_WORKSHEET_NAME]
    templated_worksheet = templated_workbook[TEMPLATED_WORKSHEET_NAME]

    for row in new_data_worksheet.iter_rows():
        for cell in row:
            templated_worksheet[cell.coordinate].value = cell.value


    combined_file_contents = io.BytesIO()
    templated_workbook.save(combined_file_contents)

    s3.put_object(
        Bucket=BUCKET,
        Key=DEST,
        Body=combined_file_contents.getvalue()
    )

    return {
        'statusCode': 200,
        'body': 'Combined file saved to S3!'
    }